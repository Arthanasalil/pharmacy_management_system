
import os
import io
import csv
import json
from datetime import datetime, timedelta, date

from flask import Blueprint, jsonify, request, send_file, abort
from sqlalchemy import (
    create_engine, text, Column, Integer, String, Float,
    DateTime, Boolean, ForeignKey, Date
)
from sqlalchemy.orm import declarative_base, sessionmaker, relationship
from dotenv import load_dotenv


try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


load_dotenv()

report_bp = Blueprint("reports", __name__)

def _build_db_url() -> str:
    
    try:
        from config import Config
        cfg = Config()
    except Exception:
        cfg = None

    host = os.getenv("MYSQL_HOST") or (cfg.MYSQL_HOST if cfg else None) or os.getenv("DB_HOST", "localhost")
    user = os.getenv("MYSQL_USER") or (cfg.MYSQL_USER if cfg else None) or os.getenv("DB_USER", "root")
    password = os.getenv("MYSQL_PASSWORD") or (cfg.MYSQL_PASSWORD if cfg else None) or os.getenv("DB_PASS", "password")
    db = os.getenv("MYSQL_DB") or (cfg.MYSQL_DB if cfg else None) or os.getenv("DB_NAME", "pharmasys")
    port = os.getenv("MYSQL_PORT") or (cfg.MYSQL_PORT if cfg else None) or os.getenv("DB_PORT", "3306")
    return f"mysql+pymysql://{user}:{password}@{host}:{port}/{db}"

DB_URL = _build_db_url()

engine = create_engine(DB_URL, echo=False, pool_pre_ping=True)
Session = sessionmaker(bind=engine)
Base = declarative_base()



class Pharmacy(Base):
    __tablename__ = "pharmacies"
    id         = Column(Integer, primary_key=True)
    code       = Column(String(20), unique=True, nullable=False)   # PHX-001
    name       = Column(String(120), nullable=False)
    location   = Column(String(200))
    is_active  = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)

    inventory  = relationship("InventoryItem", back_populates="pharmacy")
    sales      = relationship("Sale", back_populates="pharmacy")
    users      = relationship("User", back_populates="pharmacy")


class InventoryItem(Base):
    __tablename__ = "inventory_items"
    id            = Column(Integer, primary_key=True)
    pharmacy_id   = Column(Integer, ForeignKey("pharmacies.id"), nullable=False)
    drug_name     = Column(String(200), nullable=False)
    category      = Column(String(80))            # Antibiotics / Analgesics …
    batch_no      = Column(String(50))
    quantity      = Column(Integer, default=0)
    reorder_level = Column(Integer, default=50)
    unit_cost     = Column(Float, default=0.0)
    expiry_date   = Column(Date)
    updated_at    = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    pharmacy = relationship("Pharmacy", back_populates="inventory")

    @property
    def stock_status(self):
        if self.quantity == 0:
            return "out_of_stock"
        if self.quantity <= self.reorder_level:
            return "low_stock"
        return "in_stock"


class Sale(Base):
    __tablename__ = "sales"
    id            = Column(Integer, primary_key=True)
    pharmacy_id   = Column(Integer, ForeignKey("pharmacies.id"), nullable=False)
    invoice_no    = Column(String(40), unique=True)
    total_amount  = Column(Float, default=0.0)
    discount      = Column(Float, default=0.0)
    net_amount    = Column(Float, default=0.0)
    is_return     = Column(Boolean, default=False)
    sold_at       = Column(DateTime, default=datetime.utcnow)
    cashier_id    = Column(Integer, ForeignKey("users.id"))

    pharmacy = relationship("Pharmacy", back_populates="sales")


class User(Base):
    __tablename__ = "users"
    id          = Column(Integer, primary_key=True)
    pharmacy_id = Column(Integer, ForeignKey("pharmacies.id"))
    full_name   = Column(String(120), nullable=False)
    email       = Column(String(200), unique=True, nullable=False)
    role        = Column(String(40))           # admin / pharmacist / cashier …
    is_active   = Column(Boolean, default=True)
    last_login  = Column(DateTime)
    created_at  = Column(DateTime, default=datetime.utcnow)

    pharmacy    = relationship("Pharmacy", back_populates="users")


class SystemLog(Base):
    __tablename__ = "system_logs"
    id          = Column(Integer, primary_key=True)
    module      = Column(String(60))
    endpoint    = Column(String(200))
    method      = Column(String(10))
    status_code = Column(Integer)
    response_ms = Column(Float)
    user_id     = Column(Integer, ForeignKey("users.id"), nullable=True)
    logged_at   = Column(DateTime, default=datetime.utcnow)


# =============================================================================
# DB initialisation
# =============================================================================

def init_db():
    """Create tables and seed demo data if empty."""
    Base.metadata.create_all(engine)
    db = Session()
    try:
        if db.query(Pharmacy).count() == 0:
            _seed_demo_data(db)
    finally:
        db.close()


def _seed_demo_data(db: Session):
    pharmacies = [
        Pharmacy(code="PHX-001", name="MedCare Central",  location="Central District"),
        Pharmacy(code="PHX-002", name="HealthPlus North", location="North Zone"),
        Pharmacy(code="PHX-003", name="CityMed West",     location="West Block"),
        Pharmacy(code="PHX-004", name="PharmaOne South",  location="South Campus"),
    ]
    db.add_all(pharmacies)
    db.flush()

    categories = ["Antibiotics", "Analgesics", "Vitamins", "Cardiovascular"]
    import random
    random.seed(42)

    for ph in pharmacies:
        for i in range(30):
            cat = random.choice(categories)
            qty = random.randint(0, 500)
            db.add(InventoryItem(
                pharmacy_id=ph.id,
                drug_name=f"{cat} Drug #{i+1}",
                category=cat,
                batch_no=f"BT{ph.id}{i:03}",
                quantity=qty,
                reorder_level=50,
                unit_cost=round(random.uniform(5, 500), 2),
                expiry_date=date.today() + timedelta(days=random.randint(30, 730)),
            ))

        for d in range(30):
            sale_date = datetime.utcnow() - timedelta(days=d)
            for _ in range(random.randint(80, 200)):
                amt = round(random.uniform(50, 1000), 2)
                db.add(Sale(
                    pharmacy_id=ph.id,
                    invoice_no=f"INV-{ph.id}-{d}-{_}",
                    total_amount=amt,
                    net_amount=round(amt * 0.95, 2),
                    sold_at=sale_date,
                ))

    roles = ["Admin", "Pharmacist", "Cashier", "Manager"]
    names = ["Riya Sharma", "Arjun Kumar", "Priya Menon", "Vishnu Nair", "Sneha Joshi",
             "Dev Patel", "Nisha Rao", "Kiran Bose", "Amit Singh", "Lata Menon",
             "Rahul Das", "Sima Iyer"]
    for i, name in enumerate(names):
        ph = pharmacies[i % len(pharmacies)]
        db.add(User(
            pharmacy_id=ph.id,
            full_name=name,
            email=f"{name.split()[0].lower()}.{name.split()[1].lower()}@pharma.local",
            role=random.choice(roles),
            is_active=random.random() > 0.1,
            last_login=datetime.utcnow() - timedelta(hours=random.randint(0, 240)),
        ))

    db.commit()
    print("[seed] Demo data created.")


# =============================================================================
# Helpers
# =============================================================================

def _date_range(request) -> tuple[datetime, datetime]:
    """Parse ?from=YYYY-MM-DD&to=YYYY-MM-DD query params."""
    today = datetime.utcnow().date()
    from_str = request.args.get("from", str(today - timedelta(days=30)))
    to_str   = request.args.get("to",   str(today))
    try:
        date_from = datetime.strptime(from_str, "%Y-%m-%d")
        date_to   = datetime.strptime(to_str,   "%Y-%m-%d") + timedelta(hours=23, minutes=59)
    except ValueError:
        abort(400, "Invalid date format. Use YYYY-MM-DD.")
    return date_from, date_to


def _parse_date_range(from_str: str | None, to_str: str | None) -> tuple[datetime, datetime]:
    """Parse date range strings to datetime bounds."""
    today = datetime.utcnow().date()
    from_str = from_str or str(today - timedelta(days=30))
    to_str = to_str or str(today)
    try:
        date_from = datetime.strptime(from_str, "%Y-%m-%d")
        date_to = datetime.strptime(to_str, "%Y-%m-%d") + timedelta(hours=23, minutes=59)
    except ValueError:
        raise ValueError("Invalid date format. Use YYYY-MM-DD.")
    return date_from, date_to


def _pharmacy_filter(db, branch_id):
    """Return list of pharmacy IDs based on ?branch= param."""
    if branch_id and branch_id != "all":
        ph = db.query(Pharmacy).filter_by(code=branch_id).first()
        return [ph.id] if ph else []
    return [p.id for p in db.query(Pharmacy).filter_by(is_active=True).all()]


def _table_exists(table_name: str) -> bool:
    with engine.connect() as conn:
        row = conn.execute(text("SHOW TABLES LIKE :t"), {"t": table_name}).fetchone()
        return bool(row)


def _table_columns(table_name: str) -> set[str]:
    try:
        with engine.connect() as conn:
            rows = conn.execute(text(f"SHOW COLUMNS FROM {table_name}")).fetchall()
        return {str(r[0]) for r in rows}
    except Exception:
        return set()


def ensure_system_logs_table() -> bool:
    """Create system_logs table if missing. Returns True if exists/created."""
    try:
        if _table_exists("system_logs"):
            return True
        create_sql = """
            CREATE TABLE IF NOT EXISTS system_logs (
                id INT AUTO_INCREMENT PRIMARY KEY,
                module VARCHAR(60) NULL,
                endpoint VARCHAR(200) NULL,
                method VARCHAR(10) NULL,
                status_code INT NULL,
                response_ms FLOAT NULL,
                user_id INT NULL,
                logged_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """
        with engine.begin() as conn:
            conn.execute(text(create_sql))
        return True
    except Exception:
        return False


def log_system_request(module: str, endpoint: str, method: str, status_code: int, response_ms: float, user_id: int | None = None) -> None:
    """Insert a single system log row (best-effort)."""
    if not ensure_system_logs_table():
        return
    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO system_logs
                        (module, endpoint, method, status_code, response_ms, user_id, logged_at)
                    VALUES
                        (:module, :endpoint, :method, :status_code, :response_ms, :user_id, NOW())
                """),
                {
                    "module": module,
                    "endpoint": endpoint[:200],
                    "method": method[:10],
                    "status_code": int(status_code) if status_code is not None else None,
                    "response_ms": float(response_ms or 0),
                    "user_id": user_id,
                },
            )
    except Exception:
        # Best-effort logging only.
        pass


def build_inventory_report(branch: str = "all") -> dict:
    """Return per-pharmacy inventory summary using pharmacies + medicines."""
    where_branch = ""
    params = {}
    if branch and branch != "all":
        if str(branch).isdigit():
            where_branch = "AND p.pharmacyId = :branch_id"
            params["branch_id"] = int(branch)
        else:
            where_branch = "AND p.pharmacy_code = :branch"
            params["branch"] = branch

    rows = []
    try:
        if _table_exists("inventory"):
            sql = f"""
                SELECT
                    p.pharmacyId,
                    p.pharmacy_code,
                    p.name,
                    p.location,
                    COUNT(i.inventoryId) AS total_items,
                    SUM(CASE WHEN i.quantityAvailable = 0 THEN 1 ELSE 0 END) AS out_of_stock,
                    SUM(CASE WHEN i.quantityAvailable > 0 AND i.quantityAvailable <= i.minStockLevel THEN 1 ELSE 0 END) AS low_stock,
                    SUM(i.quantityAvailable * i.unitPrice) AS stock_value
                FROM pharmacies p
                LEFT JOIN inventory i
                    ON i.pharmacyId = p.pharmacyId AND i.isActive = 1
                WHERE 1=1
                {where_branch}
                GROUP BY p.pharmacyId, p.pharmacy_code, p.name, p.location
                ORDER BY p.pharmacyId
            """
            with engine.connect() as conn:
                rows = conn.execute(text(sql), params).fetchall()
        else:
            raise RuntimeError("Inventory table not found.")
    except Exception:
        if _table_exists("medicines"):
            sql = f"""
                SELECT
                    p.pharmacyId,
                    p.pharmacy_code,
                    p.name,
                    p.location,
                    COUNT(pm.medicineId) AS total_items,
                    SUM(CASE WHEN pm.stockQty = 0 THEN 1 ELSE 0 END) AS out_of_stock,
                    SUM(CASE WHEN pm.stockQty > 0 AND pm.stockQty <= pm.minStockLevel THEN 1 ELSE 0 END) AS low_stock,
                    SUM(pm.stockQty * pm.unitPrice) AS stock_value
                FROM pharmacies p
                LEFT JOIN medicines pm
                    ON pm.pharmacyId = p.pharmacyId AND pm.isActive = 1
                WHERE 1=1
                {where_branch}
                GROUP BY p.pharmacyId, p.pharmacy_code, p.name, p.location
                ORDER BY p.pharmacyId
            """
            with engine.connect() as conn:
                rows = conn.execute(text(sql), params).fetchall()
        else:
            return {"success": False, "note": "No inventory tables found.", "data": [], "total": 0}

    results = []
    for r in rows:
        total = int(r.total_items or 0)
        low_stock = int(r.low_stock or 0)
        out_of_stock = int(r.out_of_stock or 0)
        stock_value = float(r.stock_value or 0)
        fill_rate = round((total - out_of_stock) / total * 100, 1) if total else 0
        results.append({
            "pharmacy_id":   r.pharmacy_code or str(r.pharmacyId),
            "pharmacy_name": r.name,
            "location":      r.location,
            "total_items":   total,
            "low_stock":     low_stock,
            "out_of_stock":  out_of_stock,
            "stock_value":   round(stock_value, 2),
            "fill_rate":     fill_rate,
            "status": (
                "critical" if fill_rate < 70 else
                "watch"    if fill_rate < 85 else
                "healthy"
            ),
        })

    return {"success": True, "data": results, "total": len(results)}


def build_sales_report(
    branch: str = "all",
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    allow_fallback: bool = True,
) -> dict:
    """Return per-pharmacy sales summary using pharmacy_dispense_bills (fallback: bills)."""
    if not date_from or not date_to:
        date_from, date_to = _parse_date_range(None, None)

    def _select_source_table():
        if _table_exists("pharmacy_dispense_bills"):
            with engine.connect() as conn:
                cnt = conn.execute(text("SELECT COUNT(*) AS c FROM pharmacy_dispense_bills")).fetchone().c
            if cnt and int(cnt) > 0:
                return "pharmacy_dispense_bills"
        if _table_exists("bills"):
            with engine.connect() as conn:
                cnt = conn.execute(text("SELECT COUNT(*) AS c FROM bills")).fetchone().c
            if cnt and int(cnt) > 0:
                return "bills"
        return "pharmacy_dispense_bills" if _table_exists("pharmacy_dispense_bills") else ("bills" if _table_exists("bills") else None)

    source_table = _select_source_table()
    if not source_table:
        return {
            "success": False,
            "note": "No billing tables found.",
            "data": [],
            "period": {"from": str(date_from.date()), "to": str(date_to.date())},
            "totals": {"total_sales": 0, "order_count": 0, "avg_ticket": 0, "return_amount": 0.0, "net_revenue": 0.0},
        }

    def _pharmacy_ids_for_branch(branch_val):
        with engine.connect() as conn:
            if branch_val and branch_val != "all":
                if str(branch_val).isdigit():
                    return [int(branch_val)]
                row = conn.execute(
                    text("""
                        SELECT pharmacyId
                        FROM pharmacies
                        WHERE pharmacy_code = :code
                           OR name = :name
                           OR LOWER(name) = LOWER(:name)
                        LIMIT 1
                    """),
                    {"code": branch_val, "name": branch_val},
                ).fetchone()
                return [row.pharmacyId] if row else []
            rows = conn.execute(text("SELECT pharmacyId FROM pharmacies")).fetchall()
            return [r[0] for r in rows]

    def _ids_placeholders(ids):
        return ", ".join(f":ph{i}" for i in range(len(ids))) or "NULL"

    def _fetch_sales_rows(use_range: bool, ph_ids: list[int]):
        ids_sql = _ids_placeholders(ph_ids)
        where_ids = f"b.pharmacyId IN ({ids_sql})" if ph_ids else "1=1"
        has_created = "createdAt" in _table_columns(source_table)
        date_filter = "AND b.createdAt BETWEEN :df AND :dt" if (use_range and has_created) else ""
        sql = text(f"""
            SELECT
                b.pharmacyId,
                p.pharmacy_code,
                p.name,
                COUNT(b.billId) AS order_count,
                SUM(b.totalAmount) AS total_sales
            FROM {source_table} b
            LEFT JOIN pharmacies p ON p.pharmacyId = b.pharmacyId
            WHERE {where_ids}
            {date_filter}
            GROUP BY b.pharmacyId, p.pharmacy_code, p.name
            ORDER BY b.pharmacyId
        """)
        params = {f"ph{i}": ph_ids[i] for i in range(len(ph_ids))}
        if use_range and "createdAt" in _table_columns(source_table):
            params.update({"df": date_from, "dt": date_to})
        with engine.connect() as conn:
            return conn.execute(sql, params).fetchall()

    ph_ids = _pharmacy_ids_for_branch(branch)
    # Fallback if pharmacies table empty: use pharmacyIds from bills
    if not ph_ids:
        with engine.connect() as conn:
            rows = conn.execute(text("SELECT DISTINCT pharmacyId FROM pharmacy_dispense_bills")).fetchall()
            ph_ids = [int(r[0]) for r in rows if r[0] is not None]

    rows = _fetch_sales_rows(use_range=True, ph_ids=ph_ids)

    results = []
    for r in rows:
        total_sales = float(r.total_sales or 0)
        order_count = int(r.order_count or 0)
        avg_ticket = round(total_sales / order_count, 2) if order_count else 0
        results.append({
            "pharmacy_id":   r.pharmacy_code or str(r.pharmacyId),
            "pharmacy_name": r.name,
            "total_sales":   round(total_sales, 2),
            "order_count":   order_count,
            "avg_ticket":    avg_ticket,
            "return_amount": 0.0,
            "net_revenue":   round(total_sales, 2),
        })

    totals_sales = round(sum(r["total_sales"] for r in results), 2)
    totals_orders = sum(r["order_count"] for r in results)
    totals_avg = round(totals_sales / totals_orders, 2) if totals_orders else 0

    # Fallback: if last 30 days has no sales, show all-time totals
    if allow_fallback and totals_orders == 0 and totals_sales == 0:
        rows = _fetch_sales_rows(use_range=False, ph_ids=ph_ids)
        results = []
        for r in rows:
            total_sales = float(r.total_sales or 0)
            order_count = int(r.order_count or 0)
            avg_ticket = round(total_sales / order_count, 2) if order_count else 0
            results.append({
                "pharmacy_id":   r.pharmacy_code or str(r.pharmacyId),
                "pharmacy_name": r.name,
                "total_sales":   round(total_sales, 2),
                "order_count":   order_count,
                "avg_ticket":    avg_ticket,
                "return_amount": 0.0,
                "net_revenue":   round(total_sales, 2),
            })

        totals_sales = round(sum(r["total_sales"] for r in results), 2)
        totals_orders = sum(r["order_count"] for r in results)
        totals_avg = round(totals_sales / totals_orders, 2) if totals_orders else 0

        with engine.connect() as conn:
            has_created = "createdAt" in _table_columns(source_table)
            if has_created and ph_ids:
                ids_sql = _ids_placeholders(ph_ids)
                range_row = conn.execute(
                    text(f"""
                        SELECT MIN(createdAt) AS min_dt, MAX(createdAt) AS max_dt
                        FROM {source_table}
                        WHERE pharmacyId IN ({ids_sql})
                    """),
                    {f"ph{i}": ph_ids[i] for i in range(len(ph_ids))},
                ).fetchone()
            elif has_created:
                range_row = conn.execute(text(f"""
                    SELECT MIN(createdAt) AS min_dt, MAX(createdAt) AS max_dt
                    FROM {source_table}
                """)).fetchone()
            else:
                range_row = None
        period_from = range_row.min_dt.date().isoformat() if range_row and range_row.min_dt else "All time"
        period_to = range_row.max_dt.date().isoformat() if range_row and range_row.max_dt else "All time"
    else:
        period_from = str(date_from.date())
        period_to = str(date_to.date())

    return {
        "success": True,
        "data": results,
        "period": {"from": period_from, "to": period_to},
        "totals": {
            "total_sales": totals_sales,
            "order_count": totals_orders,
            "avg_ticket": totals_avg,
            "return_amount": 0.0,
            "net_revenue": totals_sales,
        },
    }


def build_user_report(role: str | None = None, branch: str = "all", status: str | None = None) -> dict:
    """Return filtered user directory with activity summary."""
    if not _table_exists("users"):
        return {"success": False, "note": "users table not found.", "summary": {"total": 0, "active": 0, "inactive": 0, "by_role": {}}, "data": []}

    where_parts = []
    params = {}
    if role:
        where_parts.append("role = :role")
        params["role"] = role
    if status:
        where_parts.append("status = :status")
        params["status"] = status

    where_clause = ""
    if where_parts:
        where_clause = "WHERE " + " AND ".join(where_parts)

    with engine.connect() as conn:
        users = conn.execute(text(f"""
            SELECT userId, name, email, role, status
            FROM users
            {where_clause}
            ORDER BY userId DESC
            LIMIT 500
        """), params).fetchall()

        totals = conn.execute(text(f"""
            SELECT role, status, COUNT(*) AS cnt
            FROM users
            {where_clause}
            GROUP BY role, status
        """), params).fetchall()

    total = sum(int(r.cnt or 0) for r in totals)
    active = sum(int(r.cnt or 0) for r in totals if (r.status or "").lower() == "approved")
    inactive = total - active
    by_role = {}
    for r in totals:
        by_role[r.role] = by_role.get(r.role, 0) + int(r.cnt or 0)

    return {
        "success": True,
        "summary": {"total": total, "active": active, "inactive": inactive, "by_role": by_role},
        "data": [
            {
                "id":          r.userId,
                "full_name":   r.name,
                "email":       r.email,
                "role":        r.role,
                "pharmacy":    None,
                "pharmacy_id": None,
                "is_active":   (r.status or "").lower() == "approved",
                "last_login":  None,
                "created_at":  None,
            }
            for r in users
        ],
    }


def build_prescription_report(date_from: datetime | None = None, date_to: datetime | None = None) -> dict:
    """Return prescription activity summary and a short daily trend."""
    if not date_from or not date_to:
        date_from, date_to = _parse_date_range(None, None)
    if not _table_exists("prescriptions"):
        return {
            "success": False,
            "note": "prescriptions table not found.",
            "period": {"from": str(date_from.date()), "to": str(date_to.date())},
            "summary": {"total": 0, "by_status": {}},
            "daily": [],
        }

    with engine.connect() as conn:
        status_rows = conn.execute(text("""
            SELECT COALESCE(status, 'unknown') AS status, COUNT(*) AS cnt
            FROM prescriptions
            WHERE createsDate BETWEEN :df AND :dt
            GROUP BY status
        """), {"df": date_from, "dt": date_to}).fetchall()

        daily_rows = conn.execute(text("""
            SELECT DATE(createsDate) AS day, COUNT(*) AS cnt
            FROM prescriptions
            WHERE createsDate >= DATE_SUB(CURDATE(), INTERVAL 6 DAY)
            GROUP BY DATE(createsDate)
            ORDER BY day
        """)).fetchall()

    by_status = {r.status: int(r.cnt or 0) for r in status_rows}
    total = sum(by_status.values())

    # Build a fixed 7-day series (today-6 .. today)
    daily_map = {str(r.day): int(r.cnt or 0) for r in daily_rows}
    today = datetime.utcnow().date()
    daily = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        daily.append({"day": day.isoformat(), "count": daily_map.get(day.isoformat(), 0)})

    return {
        "success": True,
        "period": {"from": str(date_from.date()), "to": str(date_to.date())},
        "summary": {"total": total, "by_status": by_status},
        "daily": daily,
    }


def build_system_report(hours: int = 24, module: str | None = None) -> dict:
    """Return API call counts, error rates, response times per module, server metrics."""
    since = datetime.utcnow() - timedelta(hours=hours)
    if not _table_exists("system_logs"):
        created = ensure_system_logs_table()
        server_metrics = {}
        try:
            import psutil
            server_metrics = {
                "cpu_percent":  psutil.cpu_percent(interval=0.5),
                "ram_used_gb":  round(psutil.virtual_memory().used / 1e9, 2),
                "ram_total_gb": round(psutil.virtual_memory().total / 1e9, 2),
                "ram_percent":  psutil.virtual_memory().percent,
                "disk_used_gb": round(psutil.disk_usage("/").used / 1e9, 2),
                "disk_total_gb":round(psutil.disk_usage("/").total / 1e9, 2),
                "disk_percent": psutil.disk_usage("/").percent,
            }
        except ImportError:
            server_metrics = {"note": "psutil not installed -- install for live metrics"}
        return {
            "success": True if created else False,
            "note": "No system logs yet." if created else "system_logs table not found.",
            "period_hours": hours,
            "total_api_calls": 0,
            "total_errors":    0,
            "avg_response_ms": 0,
            "error_rate":      0,
            "module_stats":    [],
            "server":          server_metrics,
            "recent":          [],
        }

    where_clause = "WHERE logged_at >= :since"
    params = {"since": since}
    if module:
        where_clause += " AND module = :module"
        params["module"] = module

    with engine.connect() as conn:
        logs = conn.execute(text(f"""
            SELECT module, endpoint, method, status_code, response_ms, logged_at
            FROM system_logs
            {where_clause}
        """), params).fetchall()

        recent_logs = conn.execute(text(f"""
            SELECT module, endpoint, method, status_code, response_ms, logged_at
            FROM system_logs
            {where_clause}
            ORDER BY logged_at DESC
            LIMIT 8
        """), params).fetchall()

    modules: dict[str, dict] = {}
    for log in logs:
        m = log.module or "unknown"
        if m not in modules:
            modules[m] = {"requests": 0, "errors": 0, "total_ms": 0.0}
        modules[m]["requests"] += 1
        if log.status_code and int(log.status_code) >= 400:
            modules[m]["errors"] += 1
        modules[m]["total_ms"] += float(log.response_ms or 0)

    module_stats = [
        {
            "module":   mod,
            "requests": d["requests"],
            "errors":   d["errors"],
            "avg_response_ms": round(d["total_ms"] / d["requests"], 1) if d["requests"] else 0,
            "error_rate": round(d["errors"] / d["requests"] * 100, 2) if d["requests"] else 0,
        }
        for mod, d in modules.items()
    ]

    server_metrics = {}
    try:
        import psutil
        server_metrics = {
            "cpu_percent":  psutil.cpu_percent(interval=0.5),
            "ram_used_gb":  round(psutil.virtual_memory().used / 1e9, 2),
            "ram_total_gb": round(psutil.virtual_memory().total / 1e9, 2),
            "ram_percent":  psutil.virtual_memory().percent,
            "disk_used_gb": round(psutil.disk_usage("/").used / 1e9, 2),
            "disk_total_gb":round(psutil.disk_usage("/").total / 1e9, 2),
            "disk_percent": psutil.disk_usage("/").percent,
        }
    except ImportError:
        server_metrics = {"note": "psutil not installed — install for live metrics"}

    total_requests = len(logs)
    total_errors = sum(m["errors"] for m in module_stats)
    avg_response = round(
        sum(m["avg_response_ms"] * m["requests"] for m in module_stats) / total_requests, 1
    ) if total_requests else 0
    error_rate = round((total_errors / total_requests) * 100, 2) if total_requests else 0

    return {
        "success": True,
        "period_hours": hours,
        "total_api_calls": total_requests,
        "total_errors":    total_errors,
        "avg_response_ms": avg_response,
        "error_rate":      error_rate,
        "module_stats":    module_stats,
        "server":          server_metrics,
        "recent": [
            {
                "module": r.module,
                "endpoint": r.endpoint,
                "method": r.method,
                "status_code": r.status_code,
                "response_ms": r.response_ms,
                "logged_at": r.logged_at.isoformat() if r.logged_at else None,
            }
            for r in recent_logs
        ],
    }


# =============================================================================
# Routes — Inventory
# =============================================================================

@report_bp.route("/api/reports/inventory", methods=["GET"])
def inventory_report():
    """
    GET /api/reports/inventory?branch=PHARM-0001&status=low_stock

    Returns per-pharmacy inventory summary using real tables:
    pharmacies + medicines.
    """
    branch = request.args.get("branch", "all")
    return jsonify(build_inventory_report(branch))


@report_bp.route("/api/reports/inventory/<string:pharmacy_code>/items", methods=["GET"])
def inventory_items(pharmacy_code: str):
    """Paginated item-level detail for one pharmacy (inventory + medicines)."""
    page     = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 25))
    offset   = (page - 1) * per_page

    with engine.connect() as conn:
        ph = conn.execute(text("""
            SELECT pharmacyId, name
            FROM pharmacies
            WHERE pharmacy_code = :code
            LIMIT 1
        """), {"code": pharmacy_code}).fetchone()

        if not ph:
            abort(404, f"Pharmacy {pharmacy_code} not found.")

        total = conn.execute(text("""
            SELECT COUNT(*) AS c
            FROM inventory i
            WHERE i.pharmacyId = :pid
              AND i.isActive = 1
        """), {"pid": ph.pharmacyId}).fetchone().c

        items = conn.execute(text("""
            SELECT
                m.name,
                m.content,
                m.brandName,
                i.quantityAvailable,
                i.unitPrice,
                i.minStockLevel,
                i.isActive,
                i.updatedAt
            FROM inventory i
            JOIN medicines m
              ON m.medicineId = i.medicineId
             AND m.pharmacyId = i.pharmacyId
            WHERE i.pharmacyId = :pid
              AND i.isActive = 1
            ORDER BY m.name
            LIMIT :limit OFFSET :offset
        """), {"pid": ph.pharmacyId, "limit": per_page, "offset": offset}).fetchall()

    def _status(qty, min_level):
        if qty == 0:
            return "out_of_stock"
        if qty <= min_level:
            return "low_stock"
        return "in_stock"

    return jsonify({
        "success": True,
        "pharmacy": ph.name,
        "page": page,
        "per_page": per_page,
        "total": int(total or 0),
        "data": [
            {
                "drug_name":  i.name,
                "category":   i.content,
                "batch_no":   i.brandName,
                "quantity":   int(i.quantityAvailable or 0),
                "unit_cost":  float(i.unitPrice or 0),
                "stock_value": round((i.quantityAvailable or 0) * (i.unitPrice or 0), 2),
                "status":     _status(int(i.quantityAvailable or 0), int(i.minStockLevel or 0)),
                "expiry":     None,
                "updated_at": i.updatedAt.isoformat() if i.updatedAt else None,
            }
            for i in items
        ],
    })


# =============================================================================
# Routes — Sales
# =============================================================================

@report_bp.route("/api/reports/sales", methods=["GET"])
def sales_report():
    """
    GET /api/reports/sales?branch=all&from=2025-03-01&to=2025-03-10

    Returns per-pharmacy revenue totals, order counts, avg ticket.
    """
    branch = request.args.get("branch", "all")
    date_from, date_to = _date_range(request)
    return jsonify(build_sales_report(branch, date_from, date_to, allow_fallback=False))

@report_bp.route("/api/reports/sales/trend", methods=["GET"])
def sales_trend():
    """
    GET /api/reports/sales/trend?branch=all&from=2025-03-01&to=2025-03-10

    Returns daily aggregated revenue â€” used for the line chart.
    """
    branch = request.args.get("branch", "all")
    date_from, date_to = _date_range(request)

    def _pharmacy_ids_for_branch(branch_val):
        with engine.connect() as conn:
            if branch_val and branch_val != "all":
                if str(branch_val).isdigit():
                    return [int(branch_val)]
                row = conn.execute(
                    text("""
                        SELECT pharmacyId
                        FROM pharmacies
                        WHERE pharmacy_code = :code
                           OR name = :name
                           OR LOWER(name) = LOWER(:name)
                        LIMIT 1
                    """),
                    {"code": branch_val, "name": branch_val},
                ).fetchone()
                return [row.pharmacyId] if row else []
            rows = conn.execute(text("SELECT pharmacyId FROM pharmacies")).fetchall()
            return [r[0] for r in rows]

    ph_ids = _pharmacy_ids_for_branch(branch)
    ph_ids_tuple = tuple(ph_ids) if ph_ids else (0,)

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT DATE(createdAt) as sale_date, SUM(totalAmount) as revenue
            FROM pharmacy_dispense_bills
            WHERE pharmacyId IN :ph_ids
              AND createdAt BETWEEN :df AND :dt
            GROUP BY DATE(createdAt)
            ORDER BY sale_date
        """), {"ph_ids": ph_ids_tuple, "df": date_from, "dt": date_to}).fetchall()

    return jsonify({
        "success": True,
        "data": [{"date": str(r[0]), "revenue": float(r[1] or 0)} for r in rows]
    })



def sales_trend():
    """
    GET /api/reports/sales/trend?branch=all&from=2025-03-01&to=2025-03-10

    Returns daily aggregated revenue — used for the line chart.
    """
    db = Session()
    try:
        branch            = request.args.get("branch", "all")
        date_from, date_to = _date_range(request)
        ph_ids             = _pharmacy_filter(db, branch)

        rows = db.execute(text("""
            SELECT DATE(sold_at) as sale_date, SUM(net_amount) as revenue
            FROM sales
            WHERE pharmacy_id IN :ph_ids
              AND sold_at BETWEEN :df AND :dt
              AND is_return = 0
            GROUP BY DATE(sold_at)
            ORDER BY sale_date
        """), {"ph_ids": tuple(ph_ids) or (0,), "df": date_from, "dt": date_to}).fetchall()

        return jsonify({
            "success": True,
            "data": [{"date": str(r[0]), "revenue": float(r[1] or 0)} for r in rows]
        })
    finally:
        db.close()


# =============================================================================
# Routes — Users
# =============================================================================

@report_bp.route("/api/reports/users", methods=["GET"])
def user_report():
    """
    GET /api/reports/users?role=Admin&branch=PHX-001&status=active

    Returns filtered user directory with activity summary.
    """
    role    = request.args.get("role")
    branch  = request.args.get("branch", "all")
    status  = request.args.get("status")   # active | inactive
    return jsonify(build_user_report(role=role, branch=branch, status=status))


# =============================================================================
# Routes — System Usage
# =============================================================================

@report_bp.route("/api/reports/system", methods=["GET"])
def system_report():
    """
    GET /api/reports/system?range=24h

    Returns API call counts, error rates, response times per module,
    and basic server metrics (CPU/RAM/disk via psutil if available).
    """
    hours  = int(request.args.get("hours", 24))
    module = request.args.get("module")
    return jsonify(build_system_report(hours=hours, module=module))

@report_bp.route("/api/reports/export", methods=["GET"])
def export_report():
    """
    GET /api/reports/export?report=inventory&format=csv|xlsx|pdf|json
                           &branch=all&from=2025-03-01&to=2025-03-10
    """
    fmt        = request.args.get("format", "csv").lower()
    report     = request.args.get("report", "inventory")
    branch     = request.args.get("branch", "all")
    date_from, date_to = _date_range(request)

    db = Session()
    try:
        ph_ids     = _pharmacy_filter(db, branch)
        pharmacies = db.query(Pharmacy).filter(Pharmacy.id.in_(ph_ids)).all()

        # Build raw rows based on report type
        if report == "inventory":
            headers = ["Pharmacy", "Code", "Total Items", "Low Stock",
                       "Out of Stock", "Stock Value (₹)", "Fill Rate %", "Status"]
            rows = []
            for ph in pharmacies:
                items        = db.query(InventoryItem).filter_by(pharmacy_id=ph.id).all()
                low          = sum(1 for i in items if i.stock_status == "low_stock")
                oos          = sum(1 for i in items if i.stock_status == "out_of_stock")
                val          = round(sum(i.quantity * i.unit_cost for i in items), 2)
                fill_rate    = round((len(items) - oos) / len(items) * 100, 1) if items else 0
                status_label = "critical" if fill_rate < 70 else "watch" if fill_rate < 85 else "healthy"
                rows.append([ph.name, ph.code, len(items), low, oos, val, fill_rate, status_label])

        elif report == "sales":
            headers = ["Pharmacy", "Code", "Total Sales (₹)", "Orders",
                       "Avg Ticket (₹)", "Returns (₹)", "Net Revenue (₹)"]
            rows = []
            for ph in pharmacies:
                sales   = db.query(Sale).filter(
                    Sale.pharmacy_id == ph.id,
                    Sale.sold_at.between(date_from, date_to),
                    Sale.is_return == False
                ).all()
                returns = db.query(Sale).filter(
                    Sale.pharmacy_id == ph.id,
                    Sale.sold_at.between(date_from, date_to),
                    Sale.is_return == True
                ).all()
                ts  = round(sum(s.total_amount for s in sales), 2)
                ret = round(sum(s.total_amount for s in returns), 2)
                rows.append([ph.name, ph.code, ts, len(sales),
                             round(ts / len(sales), 2) if sales else 0, ret,
                             round(ts - ret, 2)])

        elif report == "users":
            headers = ["Name", "Email", "Role", "Pharmacy", "Active", "Last Login"]
            users   = db.query(User).filter(User.pharmacy_id.in_(ph_ids)).all()
            rows    = [
                [u.full_name, u.email, u.role,
                 u.pharmacy.name if u.pharmacy else "",
                 "Yes" if u.is_active else "No",
                 u.last_login.strftime("%Y-%m-%d %H:%M") if u.last_login else "—"]
                for u in users
            ]
        else:
            abort(400, f"Unknown report type: {report}")

        # ── CSV ────────────────────────────────────────────────────────────────
        if fmt == "csv":
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(headers)
            writer.writerows(rows)
            buf.seek(0)
            return send_file(
                io.BytesIO(buf.getvalue().encode()),
                mimetype="text/csv",
                as_attachment=True,
                download_name=f"{report}_report_{date.today()}.csv",
            )

        # ── JSON ───────────────────────────────────────────────────────────────
        elif fmt == "json":
            payload = [dict(zip(headers, r)) for r in rows]
            return jsonify({"success": True, "report": report, "data": payload})

        # ── XLSX ───────────────────────────────────────────────────────────────
        elif fmt == "xlsx":
            if not XLSX_AVAILABLE:
                abort(500, "openpyxl not installed.")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report.capitalize()

            header_fill = PatternFill("solid", fgColor="111827")
            header_font = Font(bold=True, color="00D4AA")

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for r_idx, row in enumerate(rows, 2):
                for c_idx, val in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"{report}_report_{date.today()}.xlsx",
            )

        # ── PDF ────────────────────────────────────────────────────────────────
        elif fmt == "pdf":
            if not PDF_AVAILABLE:
                abort(500, "fpdf2 not installed.")

            pdf = FPDF(orientation="L", unit="mm", format="A4")
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 10, f"PharmaSys — {report.capitalize()} Report", ln=True, align="C")
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(0, 6, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Branch: {branch}", ln=True, align="C")
            pdf.ln(4)

            col_w = 260 // len(headers)
            pdf.set_fill_color(17, 24, 39)
            pdf.set_text_color(0, 212, 170)
            pdf.set_font("Helvetica", "B", 8)
            for h in headers:
                pdf.cell(col_w, 8, str(h), border=1, fill=True, align="C")
            pdf.ln()

            pdf.set_text_color(200, 210, 220)
            pdf.set_font("Helvetica", "", 8)
            fill = False
            for row in rows:
                pdf.set_fill_color(26, 34, 53 if fill else 17, 24, 39)
                for val in row:
                    pdf.cell(col_w, 7, str(val), border=1, align="C", fill=fill)
                pdf.ln()
                fill = not fill

            buf = io.BytesIO(pdf.output())
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"{report}_report_{date.today()}.pdf",
            )

        else:
            abort(400, f"Unsupported format: {fmt}")

    finally:
        db.close()


# =============================================================================
# Health check
# =============================================================================

@report_bp.route("/api/health", methods=["GET"])
def health():
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        db_ok = True
    except Exception as e:
        db_ok = False

    return jsonify({
        "status":    "ok" if db_ok else "degraded",
        "db":        "connected" if db_ok else "error",
        "timestamp": datetime.utcnow().isoformat(),
        "version":   "2.1.0",
    }), 200 if db_ok else 503


# =============================================================================
# Request logging middleware (writes to system_logs)
# =============================================================================

@report_bp.after_request
def log_request(response):
    """Log every API request to system_logs for the usage dashboard."""
    if not request.path.startswith("/api/"):
        return response
    if request.environ.get("system_log_recorded"):
        return response

    parts = request.path.split("/")
    module = parts[3] if len(parts) > 3 else (parts[2] if len(parts) > 2 else "unknown")
    elapsed = getattr(request, "_elapsed_ms", 0)

    log_system_request(
        module=module,
        endpoint=request.path,
        method=request.method,
        status_code=response.status_code,
        response_ms=elapsed,
        user_id=None,
    )
    request.environ["system_log_recorded"] = True

    return response


@report_bp.before_request
def start_timer():
    import time
    request._start_time = time.time()


@report_bp.after_request
def compute_elapsed(response):
    import time
    if hasattr(request, "_start_time"):
        request._elapsed_ms = round((time.time() - request._start_time) * 1000, 2)
    return response


# =============================================================================
# Entry point
# =============================================================================

if __name__ == "__main__":
    from flask import Flask
    app = Flask(__name__)
    app.register_blueprint(report_bp)
    print("🔧  Initialising database …")
    init_db()
    print("🚀  PharmaSys Report API starting on http://0.0.0.0:5000")
    app.run(host="0.0.0.0", port=5000, debug=True)

